# -*- coding: utf-8 -*-
import urllib.request, json, ssl, time, concurrent.futures, sys, os
import pandas as pd
import numpy as np
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

sys.stdout.reconfigure(encoding='utf-8')

ctx = ssl.create_default_context()
ctx.check_hostname = False
ctx.verify_mode = ssl.CERT_NONE

csv_path = r'c:\Users\Administrator\Desktop\量化策略源代码\A股全市场_8月18日突破中轴半分位选股清单.csv'
df_raw = pd.read_csv(csv_path)
print(f"★ 成功读取 302 只突破标的，开始调取官方行业与板块映射...")

def get_industry(row):
    code = str(row['code']).zfill(6)
    sym = str(row['symbol'])
    # EastMoney market id: 1 for SH (60/68), 0 for SZ (00/30), 0 for BJ (8/9/4)
    mkt = '1' if sym.startswith('sh') else '0'
    url = f'http://push2.eastmoney.com/api/qt/stock/get?secid={mkt}.{code}&fields=f57,f58,f127,f128,f116'
    req = urllib.request.Request(url, headers={'User-Agent': 'Mozilla/5.0'})
    try:
        with urllib.request.urlopen(req, context=ctx, timeout=4) as resp:
            d = json.loads(resp.read().decode('utf-8')).get('data', {})
            ind = d.get('f127', '其他')
            area = d.get('f128', '全国')
            if not ind or ind == '-':
                ind = '其他综合'
            return {'code': code, 'industry': ind, 'area': area}
    except Exception:
        return {'code': code, 'industry': '其他综合', 'area': '全国'}

ind_map = {}
with concurrent.futures.ThreadPoolExecutor(max_workers=25) as executor:
    futures = {executor.submit(get_industry, row): row['code'] for _, row in df_raw.iterrows()}
    for f in concurrent.futures.as_completed(futures):
        res = f.result()
        ind_map[res['code']] = res

# 映射到标准大类行业
def map_major_sector(ind):
    if any(k in ind for k in ['半导体', '集成电路', '电子元件', '光学光电子', '消费电子', '电子化学品', '元件', 'LED']):
        return '半导体与电子芯片'
    elif any(k in ind for k in ['软件开发', '计算机设备', '互联网服务', '通信设备', '游戏', '文化传媒', '通信服务', '影视']):
        return '计算机与AI传媒'
    elif any(k in ind for k in ['制药', '生物', '医疗', '中药', '医药', '化学制剂', '原料药', '体外诊断']):
        return '医药生物与医疗健康'
    elif any(k in ind for k in ['光伏', '电池', '风电', '电网', '储能', '电力', '能源金属', '核电']):
        return '新能源与电力储能'
    elif any(k in ind for k in ['汽车', '专用设备', '通用设备', '航空装备', '航天', '军工', '自动化', '机械', '轨交']):
        return '高端制造与机械军工'
    elif any(k in ind for k in ['证券', '银行', '保险', '多元金融', '房地产', '园区开发']):
        return '金融地产与券商'
    elif any(k in ind for k in ['工程建设', '水泥', '有色', '钢铁', '煤炭', '化工', '基础建设', '装修建材', '矿业']):
        return '大基建与周期资源'
    elif any(k in ind for k in ['食品', '饮料', '酿酒', '白酒', '家电', '商业百货', '旅游', '农牧', '纺织', '服饰']):
        return '大消费与商贸百货'
    else:
        return '轻工环保与其他'

df_raw['细分行业'] = df_raw['code'].apply(lambda c: ind_map.get(str(c).zfill(6), {}).get('industry', '其他综合'))
df_raw['所属板块'] = df_raw['细分行业'].apply(map_major_sector)
df_raw = df_raw.sort_values(by=['所属板块', 'break_pct'], ascending=[True, False]).reset_index(drop=True)

# 创建 Excel 写入器
excel_path = r'c:\Users\Administrator\Desktop\A股全市场_8月18日突破中轴半分位_分板块精选表.xlsx'
wb = openpyxl.Workbook()

# 样式定义
font_title = Font(name='微软雅黑', size=16, bold=True, color='1A365D')
font_subtitle = Font(name='微软雅黑', size=10, italic=True, color='718096')
font_sec_hdr = Font(name='微软雅黑', size=11, bold=True, color='FFFFFF')
font_hdr = Font(name='微软雅黑', size=11, bold=True, color='2D3748')
font_cell = Font(name='微软雅黑', size=10, color='2D3748')
font_bold = Font(name='微软雅黑', size=10, bold=True, color='2D3748')
font_red = Font(name='微软雅黑', size=10, bold=True, color='C53030')
font_green = Font(name='微软雅黑', size=10, bold=True, color='276749')

fill_navy = PatternFill(start_color='1A365D', end_color='1A365D', fill_type='solid')
fill_blue_hdr = PatternFill(start_color='EBF8FF', end_color='EBF8FF', fill_type='solid')
fill_zebra = PatternFill(start_color='F7FAFC', end_color='F7FAFC', fill_type='solid')
fill_highlight = PatternFill(start_color='FFF5F5', end_color='FFF5F5', fill_type='solid')

thin_border = Border(
    left=Side(style='thin', color='E2E8F0'),
    right=Side(style='thin', color='E2E8F0'),
    top=Side(style='thin', color='E2E8F0'),
    bottom=Side(style='thin', color='E2E8F0')
)

align_center = Alignment(horizontal='center', vertical='center')
align_left = Alignment(horizontal='left', vertical='center')
align_right = Alignment(horizontal='right', vertical='center')

# ==============================================================================
# Sheet 1: 📊【板块全景分布大盘】
# ==============================================================================
ws_summary = wb.active
ws_summary.title = "📊 板块全景大盘"
ws_summary.views.sheetView[0].showGridLines = True

ws_summary.cell(row=1, column=1, value="A 股全市场【8月18日突破中轴半分位】板块全景分布统计").font = font_title
ws_summary.cell(row=2, column=1, value="筛选公式: 8月18日最高价 > (6月底高点 + 7月底低点) / 2 | 全市场 5,340 只扫描命中 302 只").font = font_subtitle

sum_headers = ["序号", "核心板块名称", "达标股票数量", "板块占比", "平均突破幅度(%)", "最强突破龙头", "该板块最高突破幅度(%)"]
ws_summary.row_dimensions[4].height = 26

for col_idx, h in enumerate(sum_headers, 1):
    c = ws_summary.cell(row=4, column=col_idx, value=h)
    c.font = font_sec_hdr
    c.fill = fill_navy
    c.alignment = align_center

sector_stats = []
for sec_name, grp in df_raw.groupby('所属板块'):
    top_stock = grp.iloc[0]
    sector_stats.append({
        'sector': sec_name,
        'count': len(grp),
        'pct_of_all': len(grp) / len(df_raw) * 100.0,
        'avg_break_pct': grp['break_pct'].mean(),
        'top_name': f"{top_stock['name']} ({top_stock['code']})",
        'max_break_pct': grp['break_pct'].max()
    })

df_sec_stats = pd.DataFrame(sector_stats).sort_values(by='count', ascending=False).reset_index(drop=True)

row_start = 5
for idx, r in df_sec_stats.iterrows():
    row_num = row_start + idx
    ws_summary.row_dimensions[row_num].height = 22
    ws_summary.cell(row=row_num, column=1, value=idx+1).alignment = align_center
    ws_summary.cell(row=row_num, column=2, value=r['sector']).alignment = align_left
    ws_summary.cell(row=row_num, column=3, value=r['count']).alignment = align_center
    ws_summary.cell(row=row_num, column=4, value=f"{r['pct_of_all']:.1f}%").alignment = align_center
    
    c_avg = ws_summary.cell(row=row_num, column=5, value=f"+{r['avg_break_pct']:.2f}%")
    c_avg.alignment = align_right
    c_avg.font = font_red
    
    ws_summary.cell(row=row_num, column=6, value=r['top_name']).alignment = align_left
    c_max = ws_summary.cell(row=row_num, column=7, value=f"+{r['max_break_pct']:.2f}%")
    c_max.alignment = align_right
    c_max.font = font_red
    
    for c in range(1, 8):
        cell = ws_summary.cell(row=row_num, column=c)
        cell.border = thin_border
        if idx % 2 == 1:
            cell.fill = fill_zebra

# Total row
tot_row = row_start + len(df_sec_stats)
ws_summary.cell(row=tot_row, column=1, value="合计").alignment = align_center
ws_summary.cell(row=tot_row, column=2, value="全市场总计").alignment = align_left
ws_summary.cell(row=tot_row, column=3, value=len(df_raw)).alignment = align_center
ws_summary.cell(row=tot_row, column=4, value="100.0%").alignment = align_center
ws_summary.cell(row=tot_row, column=5, value=f"+{df_raw['break_pct'].mean():.2f}%").alignment = align_right
ws_summary.cell(row=tot_row, column=6, value=f"全市场第一: {df_raw.iloc[0]['name']} ({df_raw.iloc[0]['code']})").alignment = align_left
ws_summary.cell(row=tot_row, column=7, value=f"+{df_raw['break_pct'].max():.2f}%").alignment = align_right

for c in range(1, 8):
    cell = ws_summary.cell(row=tot_row, column=c)
    cell.font = font_bold
    cell.fill = fill_blue_hdr
    cell.border = thin_border

# ==============================================================================
# Helper Function to populate stock table
# ==============================================================================
detail_headers = ["序号", "股票代码", "股票名称", "所属板块", "细分行业", "6月底高点(元)", "7月底低点(元)", "6-7月中轴半分位(元)", "8/18最高价(元)", "突破超额(元)", "突破强度(%)"]

def write_stock_sheet(ws, df_subset, title_text):
    ws.views.sheetView[0].showGridLines = True
    ws.cell(row=1, column=1, value=title_text).font = font_title
    ws.cell(row=2, column=1, value=f"共筛选出 {len(df_subset)} 只强势突破股票 | 排序方式: 按突破强度降序排列").font = font_subtitle
    ws.row_dimensions[4].height = 24
    
    for col_idx, h in enumerate(detail_headers, 1):
        c = ws.cell(row=4, column=col_idx, value=h)
        c.font = font_sec_hdr
        c.fill = fill_navy
        c.alignment = align_center
    
    for idx, r in df_subset.iterrows():
        row_num = 5 + idx
        ws.row_dimensions[row_num].height = 20
        ws.cell(row=row_num, column=1, value=idx+1).alignment = align_center
        ws.cell(row=row_num, column=2, value=str(r['code']).zfill(6)).alignment = align_center
        ws.cell(row=row_num, column=3, value=r['name']).alignment = align_left
        ws.cell(row=row_num, column=4, value=r['所属板块']).alignment = align_center
        ws.cell(row=row_num, column=5, value=r['细分行业']).alignment = align_left
        ws.cell(row=row_num, column=6, value=f"{r['h_jun']:.2f}").alignment = align_right
        ws.cell(row=row_num, column=7, value=f"{r['l_jul']:.2f}").alignment = align_right
        ws.cell(row=row_num, column=8, value=f"{r['mid_line']:.2f}").alignment = align_right
        ws.cell(row=row_num, column=9, value=f"{r['h_aug18']:.2f}").alignment = align_right
        
        c_diff = ws.cell(row=row_num, column=10, value=f"+{r['break_diff']:.2f}")
        c_diff.alignment = align_right
        c_diff.font = font_red
        
        c_pct = ws.cell(row=row_num, column=11, value=f"+{r['break_pct']:.2f}%")
        c_pct.alignment = align_right
        c_pct.font = font_red
        
        for c in range(1, 12):
            cell = ws.cell(row=row_num, column=c)
            cell.border = thin_border
            if r['break_pct'] >= 20.0:
                cell.fill = fill_highlight
            elif idx % 2 == 1:
                cell.fill = fill_zebra

# ==============================================================================
# Sheet 2: 🏆【全市场 302 只总表】
# ==============================================================================
ws_all = wb.create_sheet(title="🏆 全市场总表(302只)")
write_stock_sheet(ws_all, df_raw.sort_values(by='break_pct', ascending=False).reset_index(drop=True), "A 股全市场【8月18日突破中轴半分位】完整选股总表")

# ==============================================================================
# Sheets 3~11: 各核心板块专属子工作表
# ==============================================================================
major_sectors_order = [
    ('💻 半导体与电子芯片', '半导体与电子芯片'),
    ('🤖 计算机与AI传媒', '计算机与AI传媒'),
    ('💊 医药生物与医疗', '医药生物与医疗健康'),
    ('⚡ 新能源与电力储能', '新能源与电力储能'),
    ('🚗 高端制造与军工', '高端制造与机械军工'),
    ('📈 金融地产与券商', '金融地产与券商'),
    ('🏗️ 大基建与周期资源', '大基建与周期资源'),
    ('🛒 大消费与商贸百货', '大消费与商贸百货'),
    ('🔬 轻工环保与其他', '轻工环保与其他')
]

for tab_title, sec_key in major_sectors_order:
    df_sec = df_raw[df_raw['所属板块'] == sec_key].sort_values(by='break_pct', ascending=False).reset_index(drop=True)
    if not df_sec.empty:
        ws_sub = wb.create_sheet(title=tab_title)
        write_stock_sheet(ws_sub, df_sec, f"【{sec_key}】8月18日突破中轴半分位精选股票池")

# Auto fit column widths for all worksheets
for ws in wb.worksheets:
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.row in [1, 2]:
                continue
            if cell.value:
                val_str = str(cell.value)
                max_len = max(max_len, len(val_str.encode('gbk', errors='ignore')))
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

wb.save(excel_path)
# Also save a copy inside workspace
repo_copy_path = r'c:\Users\Administrator\Desktop\量化策略源代码\A股全市场_8月18日突破中轴半分位_分板块精选表.xlsx'
wb.save(repo_copy_path)

print(f"★ Excel 表格生成成功！已保存至:\n  1. {excel_path}\n  2. {repo_copy_path}")
